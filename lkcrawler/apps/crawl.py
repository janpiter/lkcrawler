# -*- coding: utf-8 -*-
import os
import sys
import time
import json
from openpyxl import load_workbook
from argparse import ArgumentParser

from lkcrawler.core.controller import Controller
from lkcrawler.lib.beanstalk import Pusher, Worker
from lkcrawler.helper.config import load

reload(sys)
sys.setdefaultencoding('utf8')


def pusher(file_path=None):
    """
    :param file_path:
    :return:
    """
    try:
        config = load()
        ps = Pusher(tube=config.get('beanstalk', 'tube'),
                    host=config.get('beanstalk', 'host'),
                    port=int(config.get('beanstalk', 'port')))
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        for cells in sheet.iter_rows():
            for cell in cells:
                profile_url = cell.value
                profile_id = profile_url.split('/in/')[-1].strip('/')
                job_data = dict(url=profile_url, id=profile_id)
                ps.set_job(json.dumps(job_data))
                print job_data
    except Exception as e:
        print e


def crawler(email=None, password=None, cookies_path=None,
            result_path=None, log_path=None,
            sleep_time=None, debug=None):
    """
    :param email:
    :param password:
    :param debug:
    :return:
    """
    try:
        config = load()
        auth = dict(email=email, password=password)
        ct = Controller(auth=auth, cookies_path=cookies_path,
                        log_path=log_path, sleep_time=sleep_time,
                        debug=debug)
        w = Worker(tube=config.get('beanstalk', 'tube'),
                   worker_id=config.get('beanstalk', 'tube'),
                   host=config.get('beanstalk', 'host'),
                   port=int(config.get('beanstalk', 'port')))
        while True:
            job = w.get_job(int(config.get('beanstalk', 'timeout')))
            if job is None:
                time.sleep(30)
            else:
                job_data = json.loads(job.body)
                profile_url = job_data['url']
                profile_id = job_data['id']
                profile_file_name = '{}/{}.html'.format(result_path,
                                                        profile_id)
                if os.path.exists(profile_file_name):
                    w.buried_job(job)
                    continue
                else:
                    if not os.path.exists(result_path):
                        os.makedirs(result_path)
                    try:
                        html_data = ct.get_profile(profile_url)
                        if html_data:
                            print html_data
                            with open(profile_file_name, 'w') as opt_file:
                                opt_file.write(html_data)
                        w.delete_job(job)
                    except Exception as e:
                        print e
                        w.buried_job(job)
                        continue
    except Exception as e:
        print e


if __name__ == '__main__':
    mode = ['pusher', 'worker']
    parser = ArgumentParser(description="Linkedin profile crawler")
    parser.add_argument('mode', choices=mode)
    parser.add_argument('--file_path', help='Excel file path.')
    parser.add_argument('--email', help='Robot email address.')
    parser.add_argument('--password', help='Robot password.')
    parser.add_argument('--cookies_path', help='Robot cookies file path.',
                        default='/data/lkcrawler/cookies')
    parser.add_argument('--result_path', help='Crawler result path.',
                        default='/data/lkcrawler/result')
    parser.add_argument('--log_path', help='Crawler log path.',
                        default='/data/lkcrawler/log')
    parser.add_argument('--sleep_time', default="5")
    parser.add_argument('--debug', choices=['True', 'False'],
                        default='False')

    args = parser.parse_args()
    _mode = args.mode
    _file_path = args.file_path
    _email = args.email
    _password = args.password
    _cookies_path = args.cookies_path
    _result_path = args.result_path
    _log_path = args.log_path
    _sleep_time = float(args.sleep_time)
    _debug = eval(args.debug)

    if _mode == 'pusher':
        pusher(file_path=_file_path)
    elif _mode == 'worker':
        if _email and _password:
            crawler(email=_email, password=_password,
                    cookies_path=_cookies_path,
                    result_path=_result_path,
                    log_path=_log_path,
                    sleep_time=_sleep_time,
                    debug=_debug)
    else:
        parser.print_help()
